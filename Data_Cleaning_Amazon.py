#!/usr/bin/env python
# coding: utf-8

# In[1]:


import pandas as pd
import numpy as np
import datetime


# In[18]:


x=pd.datetime.today()
current_date = x.strftime('%d-%m-%Y')
current_date


# In[3]:


day=pd.datetime.today()
day=day.strftime('%d')
day


# In[4]:


url1="C:/Users/SurajC/Downloads/Desktop/amazon/{}.json".format(day)
json_data=pd.read_json(url1)
url2="C:/Users/SurajC/Downloads/Desktop/amazon/{}.csv".format(day)
csv_data=pd.read_csv(url2)

json_data = pd.DataFrame(json_data)
csv_data = pd.DataFrame(csv_data)


# In[5]:


all_data = pd.concat([json_data, csv_data])
all_data = all_data.reset_index(drop=True)
all_data.drop_duplicates(inplace=True)
all_data


# In[6]:


all_data['date'] = pd.to_datetime(all_data['date'])
all_data['date'] = all_data['date'].dt.date
all_data['product_rating'] = all_data['product_rating'].str.replace(' out of 5 stars', '')
all_data['tyre_size'] = all_data['tyre_size'].str.replace('"', '')
all_data['tyre_size'] = all_data['tyre_size'].str.replace(' tyre', '')
all_data['product_mrp'] = all_data['product_mrp'].str.replace('₹', '')
all_data['sale_price'] = all_data['sale_price'].str.replace(',', '')
all_data['product_mrp'] = all_data['product_mrp'].str.replace(',', '')
all_data = all_data.fillna(value=np.nan)
all_data['brand'] = all_data['product_name'].str.split().str[0]
all_data = all_data[['product_name', 'sale_price', 'product_mrp', 'product_rating',
       'total_ratings', 'tyre_size', 'brand', 'date']]
all_data['brand'] = all_data['brand'].str.lower()
all_data.replace({'brand': {'yu':'yu wheels','mahavir':'michelin','ling':'ling long','invalid':'mrf','apolloapterra':'apollo','ralco,':'ralco','ultra':'ultramile'}},inplace=True)
all_data.drop('tyre_size',axis=1,inplace=True)
all_data.head(10)


# In[7]:


a=all_data['product_name']
a=pd.DataFrame(a)
a
a['product_name'] = a['product_name'].str.replace(' R', 'R')
a['product_name'] = a['product_name'].str.replace('P','P ')
a['product_name'] = a['product_name'].str.replace(' r', 'r')
a['product_name'] = a['product_name'].str.replace('p','p ')
a


# In[8]:


import re
b=[]
str=a['product_name']
for i in str:
    
    if re.findall(r' (\d{3}/)', i):
        x=re.findall(r' (\d{3}/)', i)
        if x[0] in i:
            start_index = i.index(x[0])
            s= i[start_index:start_index+10]
            b.append(s)
            
    else :
        b.append('--')
    

df1 = pd.DataFrame(b)
a['tyre_size'] = pd.DataFrame(b)
print(a)


# In[9]:


a['tyre_size'] = a['tyre_size'].str.replace(',', '')
a['tyre_size'] = a['tyre_size'].str.replace('215/75/15/', '215/75/15')
a['tyre_size'] = a['tyre_size'].str.replace('165/70/14/', '165/70/14')
a['tyre_size'] = a['tyre_size'].str.replace('235/70/16/', '235/70/16')
a['tyre_size'] = a['tyre_size'].str.replace('215/75/15/', '215/75/15')
a['tyre_size'] = a['tyre_size'].str.replace('165/70/14/', '165/70/14')
a['tyre_size'] = a['tyre_size'].str.replace('235/70/16/', '235/70/16')
a['tyre_size'] = a['tyre_size'].str.replace('215/75/15/', '215/75/15')
a['tyre_size'] = a['tyre_size'].str.replace('235/70/16/', '235/70/16')
a['tyre_size'] = a['tyre_size'].str.replace('215/65/16/', '215/65/16')
a['tyre_size'] = a['tyre_size'].str.replace('235/70/16/', '235/70/16')
a['tyre_size'] = a['tyre_size'].str.replace('195/55/16/', '195/55/16')
a['tyre_size'] = a['tyre_size'].str.replace('175/65/15/', '175/65/15')
a['tyre_size'] = a['tyre_size'].str.replace('110/70/17/', '110/70/17')
a['tyre_size'] = a['tyre_size'].str.replace('185/80D14', '185/80/14')
a['tyre_size'] = a['tyre_size'].str.replace('215/60/17/', '215/60/17')
a['tyre_size'] = a['tyre_size'].str.replace('145/70/80/', '145/70/80')
a['tyre_size'] = a['tyre_size'].str.replace('215/60/16/', '215/60/16')
a['tyre_size'] = a['tyre_size'].str.replace('215/65/16/', '215/65/16')
a['tyre_size'] = a['tyre_size'].str.replace('215/75/15/', '215/75/15')
a['tyre_size'] = a['tyre_size'].str.replace('102/1008', '102/10/08')
a['tyre_size'] = a['tyre_size'].str.replace('165/80D13', '165/80/13')
a['tyre_size'] = a['tyre_size'].str.replace('155/80/D', '155/80')
a['tyre_size'] = a['tyre_size'].str.replace('235/6517/', '235/65/17')
a['tyre_size'] = a['tyre_size'].str.replace('205/65 R16', '205/65/16')
a['tyre_size'] = a['tyre_size'].str.replace('215/60R17-', '215/60/17')
a['tyre_size'] = a['tyre_size'].str.replace('165/70R14 ', '165/70/14')
a['tyre_size'] = a['tyre_size'].str.replace('205/65R16', '205/65/16')
a['tyre_size'] = a['tyre_size'].str.replace('155/70R13', '155/70/13')
a['tyre_size'] = a['tyre_size'].str.replace('ST', '')
a['tyre_size'] = a['tyre_size'].str.replace('165/70/14/', '165/70/14')
a['tyre_size'] = a['tyre_size'].str.replace('235/65/', '235/65')
a['tyre_size'] = a['tyre_size'].str.replace('LT', '')
a['tyre_size'] = a['tyre_size'].str.replace('265/60/18/', '265/60/18')
a['tyre_size'] = a['tyre_size'].str.replace('235/75/15/', '235/75/15')
a['tyre_size'] = a['tyre_size'].str.replace('(', '')
a['tyre_size'] = a['tyre_size'].str.replace('145/70/12/', '145/70/12')
a['tyre_size'] = a['tyre_size'].str.replace('145/55/70/', '145/55/70')
a['tyre_size'] = a['tyre_size'].str.replace('205/60/16/', '205/60/16')
a['tyre_size'] = a['tyre_size'].str.replace('205/60/16/', '205/60/16')
a['tyre_size'] = a['tyre_size'].str.replace('145/70/12/', '145/70/12')
a['tyre_size'] = a['tyre_size'].str.replace(' ', '')
a['tyre_size'] = a['tyre_size'].str.replace('R', '/')
a['tyre_size'] = a['tyre_size'].str.replace('-', '/')
a['tyre_size'] = a['tyre_size'].str.replace('//', '/')
a['tyre_size'] = a['tyre_size'].str.replace('r', '/')
a['tyre_size'] = a['tyre_size'].str.replace('A', '')
a['tyre_size'] = a['tyre_size'].str.replace('z', '')
a['tyre_size'] = a['tyre_size'].str.replace('k', '')
a['tyre_size'] = a['tyre_size'].str.replace('K', '')
a['tyre_size'] = a['tyre_size'].str.replace('Z', '')
a['tyre_size'] = a['tyre_size'].str.replace('205/60/16/', '205/60/16')
a['tyre_size'] = a['tyre_size'].str.replace('215/75/15/', '215/75/15')
a['tyre_size'].unique()
a


# In[10]:


all_data['tyre_size']=a['tyre_size']
all_data['brand']=all_data['brand'].str.upper()
all_data


# In[11]:


all_data.fillna(0,inplace=True)
all_data.replace(np.NaN, 0,inplace=True)
convert_dict = {'sale_price': 'int',
                'product_mrp': 'float',
                'product_rating':'float',
                'total_ratings':'int'
               }

all_data = all_data.astype(convert_dict)
all_data.dtypes


# In[12]:


all_data.drop_duplicates(inplace=True)
all_data.to_excel('C:/Users/SurajC/Downloads/Desktop/amazon/{}.xlsx'.format(current_date),index=False)
all_data.to_excel('C:/Users/SurajC/Downloads/Desktop/bridgestone-dashboard/app/reports/{}.xlsx'.format(current_date),index=False)
all_data.shape


# In[13]:


from openpyxl import load_workbook
# new dataframe with same columns
writer = pd.ExcelWriter('C:/Users/SurajC/Downloads/Desktop/unique.xlsx', engine='openpyxl')
# try to open an existing workbook
writer.book = load_workbook('C:/Users/SurajC/Downloads/Desktop/unique.xlsx')
# copy existing sheets
writer.sheets = dict((ws.title, ws) for ws in writer.book.worksheets)
# read existing file
reader = pd.read_excel(r'C:/Users/SurajC/Downloads/Desktop/unique.xlsx')
# write out the new sheet
all_data.to_excel(writer,index=False,header=False,startrow=len(reader)+1)

writer.close()


# In[14]:


selection = ['145/80/12','165/80/14','155/70/13','175/65/14','155/80/13','185/65/15','215/75/15',
'155/65/13','205/65/15','185/70/14','145/80/13','145/70/12','145/70/13','165/70/14','165/65/14',
'165/65/13','185/65/14','185/60/15','195/55/16','175/70/13','155/65/14','175/70/14','235/70/16',
'205/65/16','175/65/15','205/60/16','235/65/17','215/60/16','195/65/15','195/60/15','215/65/16']
df = all_data[pd.DataFrame(all_data.tyre_size.tolist()).isin(selection).any(1).values]

df.shape


# In[15]:


selection = ['APOLLO','CEAT','ULTRA MILE','JK','BRIDGESTONE','MICHELIN','GOODYEAR','MRF','ULTRAMILE','YOKOHAMA']
df1 = df[pd.DataFrame(df.brand.tolist()).isin(selection).any(1).values]
df1.drop_duplicates()
df1


# In[16]:


# df1.to_excel('C:/Users/SurajC/Downloads/Desktop/amazon/32_tyre_size.xlsx',index=False)


# In[17]:


from openpyxl import load_workbook
# new dataframe with same columns
writer = pd.ExcelWriter('C:/Users/SurajC/Downloads/Desktop/32_tyre_size.xlsx', engine='openpyxl')
# try to open an existing workbook
writer.book = load_workbook('C:/Users/SurajC/Downloads/Desktop/32_tyre_size.xlsx')
# copy existing sheets
writer.sheets = dict((ws.title, ws) for ws in writer.book.worksheets)
# read existing file
reader = pd.read_excel(r'C:/Users/SurajC/Downloads/Desktop/32_tyre_size.xlsx')
# write out the new sheet
df1.to_excel(writer,index=False,header=False,startrow=len(reader)+1)

writer.close()


# In[ ]:




